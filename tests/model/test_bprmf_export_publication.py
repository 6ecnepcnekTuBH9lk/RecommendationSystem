import csv
import json
import threading
from concurrent.futures import ThreadPoolExecutor
from pathlib import Path

import pandas as pd
import pytest
import torch
from openpyxl import load_workbook

from Application.model import BPRMF


class _SyntheticExportModel:
    def __init__(self, inference_error=None):
        self.inference_error = inference_error

    def item_vec_all(self):
        return torch.tensor([[1.0]], dtype=torch.float32)

    def user_emb(self, user_indices):
        if self.inference_error is not None:
            raise self.inference_error
        return torch.ones((len(user_indices), 1), dtype=torch.float32)


class _SyntheticRecommendationImpactModel:
    def item_vec_all(self):
        return torch.tensor([[2.0], [1.0]], dtype=torch.float32)

    def user_emb(self, user_indices):
        return torch.ones((len(user_indices), 1), dtype=torch.float32)


def _prepare_synthetic_export(
    tmp_path,
    monkeypatch,
    *,
    model=None,
    train_item_meta=None,
    idx2item=None,
):
    monkeypatch.chdir(tmp_path)
    data_dir = tmp_path / "synthetic-data"
    data_dir.mkdir()
    pd.DataFrame(
        [
            {
                "MindboxID": "user-1",
                "КодНоменклатуры": "item-1",
                "Количество": "1",
                "Телефон": "+7 (900) 123-45-67",
                "ДисконтнаяКарта": "card-1",
                "Почта": "user-1@example.test",
            }
        ]
    ).to_csv(
        data_dir / "Заказы.csv",
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )
    pd.DataFrame(
        columns=["MindboxID", "КодНоменклатуры", "ТипТовара"]
    ).to_csv(
        data_dir / "Просмотры.csv",
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )
    pd.DataFrame(columns=["MindboxID", "КодНоменклатуры"]).to_csv(
        data_dir / "Избранное.csv",
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )

    cfg = BPRMF.TrainConfig(data_dir=str(data_dir), topk=1)
    mappings = {
        "idx2user": ["user-1"],
        "idx2item": idx2item or ["item-1"],
    }
    checkpoint = {"train_item_meta": train_item_meta or {}}
    export_model = model or _SyntheticExportModel()

    monkeypatch.setattr(
        BPRMF,
        "_load_artifacts",
        lambda model_dir="Модель": (mappings, checkpoint),
    )
    monkeypatch.setattr(
        BPRMF,
        "_build_model_from_ckpt",
        lambda loaded_checkpoint, device: (
            export_model,
            cfg,
            1,
            len(mappings["idx2item"]),
        ),
    )
    monkeypatch.setattr(
        BPRMF,
        "_load_selected_collections_from_settings",
        lambda: [],
    )
    monkeypatch.setattr(BPRMF, "_load_item_names", lambda data_dir: {"item-1": "Synthetic item"})
    monkeypatch.setattr(BPRMF, "_load_item_stocks", lambda data_dir: {"item-1": "150"})
    monkeypatch.setattr(
        BPRMF,
        "_load_historical_item_conversion",
        lambda **kwargs: ({"item-1": 12.34}, 12.34),
    )

    output_dir = tmp_path / "outputs"
    output_dir.mkdir()
    return {
        "xlsx": output_dir / "Рекомендации.xlsx",
        "csv1": output_dir / "InternetMagazin.csv",
        "csv2": output_dir / "Mindbox.csv",
    }


def _run_export(paths, **overrides):
    arguments = {
        "out_xlsx": str(paths["xlsx"]),
        "out_csv_format1": str(paths["csv1"]) if paths["csv1"] else None,
        "out_csv_kanzler_ml": str(paths["csv2"]) if paths["csv2"] else None,
        "k": 1,
        "filter_seen": False,
        "max_export_users": 1,
        "device_str": "cpu",
    }
    arguments.update(overrides)
    return BPRMF.export_recommendations_excel(
        **arguments,
    )


def _write_old_outputs(paths):
    old_bytes = {
        "xlsx": b"old-complete-xlsx",
        "csv1": b"old-complete-internet-csv",
        "csv2": b"old-complete-mindbox-csv",
    }
    for name, path in paths.items():
        if path is not None:
            Path(path).parent.mkdir(parents=True, exist_ok=True)
            Path(path).write_bytes(old_bytes[name])
    return old_bytes


def _assert_old_outputs(paths, old_bytes, names=("xlsx", "csv1", "csv2")):
    for name in names:
        assert Path(paths[name]).read_bytes() == old_bytes[name]


def _assert_no_export_temps(root):
    assert list(Path(root).rglob("*.tmp")) == []


def _read_csv_rows(path):
    raw = Path(path).read_bytes()
    assert raw.startswith(b"\xef\xbb\xbf")
    assert b"\r\n" in raw
    assert b"\n" not in raw.replace(b"\r\n", b"")
    with Path(path).open("r", encoding="utf-8-sig", newline="") as file_object:
        return list(csv.reader(file_object, delimiter=";"))


def _read_xlsx_rows(path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        return list(workbook["Рекомендации"].iter_rows(values_only=True))
    finally:
        workbook.close()


def _sequential_name_loader(results):
    calls = []

    def load(data_dir):
        calls.append(Path(data_dir).resolve())
        result = results[len(calls) - 1]
        if isinstance(result, BaseException):
            raise result
        return result

    return load, calls


@pytest.mark.parametrize(
    "missing_filename",
    ["Заказы.csv", "Просмотры.csv", "Избранное.csv"],
)
def test_export_missing_required_interaction_source_aborts_and_preserves_outputs(
    tmp_path,
    monkeypatch,
    missing_filename,
):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)
    (tmp_path / "synthetic-data" / missing_filename).unlink()

    with pytest.raises(FileNotFoundError) as exc_info:
        _run_export(paths)

    assert missing_filename in str(exc_info.value)
    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


@pytest.mark.parametrize("empty_filename", ["Просмотры.csv", "Избранное.csv"])
def test_export_accepts_schema_valid_empty_required_interaction_source(
    tmp_path,
    monkeypatch,
    empty_filename,
):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    source_path = tmp_path / "synthetic-data" / empty_filename

    assert source_path.is_file()
    assert pd.read_csv(source_path, sep="|", dtype=str).empty
    assert _run_export(paths) == str(paths["xlsx"])


def test_missing_views_does_not_publish_incomplete_seen_recommendation(
    tmp_path,
    monkeypatch,
):
    paths = _prepare_synthetic_export(
        tmp_path,
        monkeypatch,
        model=_SyntheticRecommendationImpactModel(),
        idx2item=["item-a", "item-b"],
    )
    data_dir = tmp_path / "synthetic-data"
    pd.DataFrame(
        [
            {
                "MindboxID": "user-1",
                "КодНоменклатуры": "item-a",
                "ТипТовара": "Номенклатура",
            }
        ]
    ).to_csv(
        data_dir / "Просмотры.csv",
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )
    monkeypatch.setattr(
        BPRMF,
        "_load_item_names",
        lambda source_dir: {"item-a": "Item A", "item-b": "Item B"},
    )
    monkeypatch.setattr(
        BPRMF,
        "_load_item_stocks",
        lambda source_dir: {"item-a": "150", "item-b": "150"},
    )

    assert _run_export(paths, filter_seen=True) == str(paths["xlsx"])
    assert _read_xlsx_rows(paths["xlsx"])[1][4] == "item-b"
    published_bytes = {
        name: Path(path).read_bytes()
        for name, path in paths.items()
    }

    (data_dir / "Просмотры.csv").unlink()
    with pytest.raises(FileNotFoundError) as exc_info:
        _run_export(paths, filter_seen=True)

    assert "Просмотры.csv" in str(exc_info.value)
    _assert_old_outputs(paths, published_bytes)
    assert _read_xlsx_rows(paths["xlsx"])[1][4] == "item-b"
    _assert_no_export_temps(tmp_path)


def test_inference_error_keeps_existing_outputs_byte_for_byte(
    tmp_path, monkeypatch
):
    paths = _prepare_synthetic_export(
        tmp_path,
        monkeypatch,
        model=_SyntheticExportModel(
            inference_error=RuntimeError("synthetic inference error")
        ),
    )
    old_bytes = _write_old_outputs(paths)

    with pytest.raises(RuntimeError, match="synthetic inference error"):
        _run_export(paths)

    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


def test_export_keeps_first_names_when_second_load_fails(
    tmp_path, monkeypatch, capsys
):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    loader, calls = _sequential_name_loader([
        {"item-1": "First successful name"},
        PermissionError("synthetic second item names failure"),
    ])
    monkeypatch.setattr(BPRMF, "_load_item_names", loader)

    assert _run_export(paths) == str(paths["xlsx"])

    assert len(calls) == 2
    assert _read_xlsx_rows(paths["xlsx"])[1][5] == "First successful name"
    diagnostic = capsys.readouterr().err
    assert diagnostic.count("Не удалось загрузить названия товаров") == 1
    assert "synthetic second item names failure" in diagnostic


def test_export_uses_second_names_when_first_load_fails(
    tmp_path, monkeypatch, capsys
):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    loader, calls = _sequential_name_loader([
        PermissionError("synthetic first item names failure"),
        {"item-1": "Second successful name"},
    ])
    monkeypatch.setattr(BPRMF, "_load_item_names", loader)

    assert _run_export(paths) == str(paths["xlsx"])

    assert len(calls) == 2
    assert _read_xlsx_rows(paths["xlsx"])[1][5] == "Second successful name"
    diagnostic = capsys.readouterr().err
    assert diagnostic.count("Не удалось загрузить названия товаров") == 1
    assert "synthetic first item names failure" in diagnostic


def test_export_both_name_loads_fail_once_and_use_training_fallback(
    tmp_path, monkeypatch, capsys
):
    paths = _prepare_synthetic_export(
        tmp_path,
        monkeypatch,
        train_item_meta={
            "item-1": {"НазваниеНаСайте": "Training fallback name"}
        },
    )
    valid_loader, _calls = _sequential_name_loader([
        {"item-1": "Baseline item name"},
        {"item-1": "Baseline item name"},
    ])
    monkeypatch.setattr(BPRMF, "_load_item_names", valid_loader)
    assert _run_export(paths) == str(paths["xlsx"])
    baseline_csv1 = paths["csv1"].read_bytes()
    baseline_csv2 = paths["csv2"].read_bytes()
    capsys.readouterr()

    failing_loader, calls = _sequential_name_loader([
        PermissionError("synthetic first item names failure"),
        pd.errors.ParserError("synthetic second item names failure"),
    ])
    monkeypatch.setattr(BPRMF, "_load_item_names", failing_loader)
    assert _run_export(paths) == str(paths["xlsx"])

    assert len(calls) == 2
    assert paths["csv1"].read_bytes() == baseline_csv1
    assert paths["csv2"].read_bytes() == baseline_csv2
    assert _read_xlsx_rows(paths["xlsx"])[1][5] == "Training fallback name"
    diagnostic = capsys.readouterr().err
    assert diagnostic.count("Не удалось загрузить названия товаров") == 1
    assert "synthetic first item names failure" in diagnostic


def test_export_name_failure_without_fallback_keeps_row_and_empty_name(
    tmp_path, monkeypatch, capsys
):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    failing_loader, _calls = _sequential_name_loader([
        PermissionError("synthetic item names failure"),
        PermissionError("synthetic item names failure"),
    ])
    monkeypatch.setattr(BPRMF, "_load_item_names", failing_loader)

    assert _run_export(paths) == str(paths["xlsx"])

    row = _read_xlsx_rows(paths["xlsx"])[1]
    assert row[4] == "item-1"
    assert row[5] is None
    assert _read_csv_rows(paths["csv1"])[1] == ["79001234567", "item-1"]
    assert _read_csv_rows(paths["csv2"])[1][2] == "item-1"
    assert capsys.readouterr().err.count(
        "Не удалось загрузить названия товаров"
    ) == 1


def test_export_without_item_names_does_not_call_loader(tmp_path, monkeypatch):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)

    def fail_if_called(data_dir):
        raise AssertionError("item names loader must not be called")

    monkeypatch.setattr(BPRMF, "_load_item_names", fail_if_called)

    assert _run_export(paths, include_item_names=False) == str(paths["xlsx"])

    rows = _read_xlsx_rows(paths["xlsx"])
    assert "НазваниеНоменклатуры_1" not in rows[0]
    assert _read_csv_rows(paths["csv1"])[1] == ["79001234567", "item-1"]


def test_export_name_failure_uses_current_name_for_seasonally_mapped_item(
    tmp_path, monkeypatch, capsys
):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    cfg = BPRMF.TrainConfig(
        data_dir=str(tmp_path / "synthetic-data"),
        topk=1,
    )
    mappings = {
        "idx2user": ["user-1"],
        "idx2item": ["old-item"],
    }
    checkpoint = {
        "train_item_meta": {
            "old-item": {
                "Коллекция": "Весна-Лето 2025",
                "ВидНоменклатуры": "Jacket",
                "НазваниеНаСайте": "Old item name",
            }
        }
    }
    monkeypatch.setattr(
        BPRMF,
        "_load_artifacts",
        lambda model_dir="Модель": (mappings, checkpoint),
    )
    monkeypatch.setattr(
        BPRMF,
        "_build_model_from_ckpt",
        lambda loaded_checkpoint, device: (
            _SyntheticExportModel(),
            cfg,
            1,
            1,
        ),
    )
    monkeypatch.setattr(
        BPRMF,
        "_load_selected_collections_from_settings",
        lambda: ["Весна-Лето 2026"],
    )
    current_dir = tmp_path / "ВходныеДанные"
    current_dir.mkdir()
    pd.DataFrame(
        [
            {
                "КодНоменклатуры": "new-item",
                "Коллекция": "Весна-Лето 2026",
                "ВидНоменклатуры": "Jacket",
                "НазваниеНаСайте": "Current mapped name",
                "Остаток": "150",
            }
        ]
    ).to_csv(
        current_dir / "Номенклатура.csv",
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )
    monkeypatch.setattr(
        BPRMF,
        "_load_item_stocks",
        lambda data_dir: {"new-item": "150"},
    )
    failing_loader, _calls = _sequential_name_loader([
        PermissionError("synthetic item names failure"),
        PermissionError("synthetic item names failure"),
    ])
    monkeypatch.setattr(BPRMF, "_load_item_names", failing_loader)

    assert _run_export(paths) == str(paths["xlsx"])

    row = _read_xlsx_rows(paths["xlsx"])[1]
    assert row[4] == "new-item"
    assert row[5] == "Current mapped name"
    assert capsys.readouterr().err.count(
        "Не удалось загрузить названия товаров"
    ) == 1


def test_broken_selected_collection_settings_preserve_existing_outputs(
    tmp_path, monkeypatch
):
    real_settings_loader = BPRMF._load_selected_collections_from_settings
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)
    settings_dir = tmp_path / "Настройки"
    settings_dir.mkdir()
    (settings_dir / "filter_settings.json").write_text(
        '{"seasons_selected": [',
        encoding="utf-8",
    )
    monkeypatch.setattr(
        BPRMF,
        "_load_selected_collections_from_settings",
        real_settings_loader,
    )

    with pytest.raises(json.JSONDecodeError):
        _run_export(paths)

    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


def test_missing_selected_collection_settings_still_allow_export(
    tmp_path, monkeypatch
):
    real_settings_loader = BPRMF._load_selected_collections_from_settings
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    monkeypatch.setattr(
        BPRMF,
        "_load_selected_collections_from_settings",
        real_settings_loader,
    )

    assert _run_export(paths) == str(paths["xlsx"])


def test_valid_empty_selected_collection_settings_still_allow_export(
    tmp_path, monkeypatch
):
    real_settings_loader = BPRMF._load_selected_collections_from_settings
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    settings_dir = tmp_path / "Настройки"
    settings_dir.mkdir()
    (settings_dir / "filter_settings.json").write_text(
        json.dumps({"seasons_selected": []}),
        encoding="utf-8",
    )
    monkeypatch.setattr(
        BPRMF,
        "_load_selected_collections_from_settings",
        real_settings_loader,
    )

    assert _run_export(paths) == str(paths["xlsx"])


def test_current_stock_read_error_does_not_fallback_or_publish(
    tmp_path, monkeypatch
):
    real_stock_loader = BPRMF._load_item_stocks
    real_read_csv_pipe = BPRMF._read_csv_pipe
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)

    current_data_dir = tmp_path / "ВходныеДанные"
    current_data_dir.mkdir()
    current_path = current_data_dir / "Номенклатура.csv"
    historical_path = tmp_path / "synthetic-data" / "Номенклатура.csv"
    pd.DataFrame(
        [{"КодНоменклатуры": "item-1", "Остаток": "200"}]
    ).to_csv(
        current_path,
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )
    pd.DataFrame(
        [{"КодНоменклатуры": "item-1", "Остаток": "150"}]
    ).to_csv(
        historical_path,
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )

    current_reads = 0
    stock_load_paths = []

    def fail_second_current_read(path):
        nonlocal current_reads
        if Path(path).resolve() == current_path.resolve():
            current_reads += 1
            if current_reads == 2:
                raise PermissionError("synthetic current stock read failure")
        return real_read_csv_pipe(path)

    def recording_stock_loader(data_dir):
        stock_load_paths.append(Path(data_dir).resolve())
        return real_stock_loader(data_dir)

    monkeypatch.setattr(BPRMF, "_read_csv_pipe", fail_second_current_read)
    monkeypatch.setattr(BPRMF, "_load_item_stocks", recording_stock_loader)

    with pytest.raises(
        PermissionError,
        match="synthetic current stock read failure",
    ):
        _run_export(paths)

    assert current_reads == 2
    assert stock_load_paths == [current_data_dir.resolve()]
    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


def test_valid_empty_current_stocks_do_not_use_historical_fallback(
    tmp_path, monkeypatch
):
    real_stock_loader = BPRMF._load_item_stocks
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)

    current_data_dir = tmp_path / "ВходныеДанные"
    current_data_dir.mkdir()
    pd.DataFrame(columns=["КодНоменклатуры", "Остаток"]).to_csv(
        current_data_dir / "Номенклатура.csv",
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )
    pd.DataFrame(
        [{"КодНоменклатуры": "item-1", "Остаток": "150"}]
    ).to_csv(
        tmp_path / "synthetic-data" / "Номенклатура.csv",
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )

    stock_load_dirs = []

    def recording_stock_loader(data_dir):
        stock_load_dirs.append(Path(data_dir).resolve())
        return real_stock_loader(data_dir)

    monkeypatch.setattr(BPRMF, "_load_item_stocks", recording_stock_loader)

    assert _run_export(paths) == str(paths["xlsx"])

    assert stock_load_dirs == [current_data_dir.resolve()]
    assert _read_csv_rows(paths["csv1"]) == [["CustomerID", "ProductID"]]
    assert _read_csv_rows(paths["csv2"]) == [[
        "CustomerMindboxId",
        "Quantity",
        "ProductGroupOffline1C",
        "CustomFieldKoefficient",
    ]]


def test_missing_current_stocks_use_historical_fallback(tmp_path, monkeypatch):
    real_stock_loader = BPRMF._load_item_stocks
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    historical_data_dir = tmp_path / "synthetic-data"
    pd.DataFrame(
        [{"КодНоменклатуры": "item-1", "Остаток": "150"}]
    ).to_csv(
        historical_data_dir / "Номенклатура.csv",
        sep="|",
        index=False,
        encoding="utf-8-sig",
    )

    stock_load_dirs = []

    def recording_stock_loader(data_dir):
        stock_load_dirs.append(Path(data_dir).resolve())
        return real_stock_loader(data_dir)

    monkeypatch.setattr(BPRMF, "_load_item_stocks", recording_stock_loader)

    assert _run_export(paths) == str(paths["xlsx"])

    assert stock_load_dirs == [historical_data_dir.resolve()]
    assert _read_csv_rows(paths["csv1"])[1] == ["79001234567", "item-1"]


def test_broken_historical_stock_fallback_propagates_and_preserves_outputs(
    tmp_path, monkeypatch
):
    real_stock_loader = BPRMF._load_item_stocks
    real_read_csv_pipe = BPRMF._read_csv_pipe
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)
    historical_path = tmp_path / "synthetic-data" / "Номенклатура.csv"
    historical_path.write_text(
        "КодНоменклатуры|Остаток\nitem-1|150\n",
        encoding="utf-8-sig",
    )
    stock_load_dirs = []

    def fail_historical_stock_read(path):
        if Path(path).resolve() == historical_path.resolve():
            raise PermissionError("synthetic historical stock read failure")
        return real_read_csv_pipe(path)

    def recording_stock_loader(data_dir):
        stock_load_dirs.append(Path(data_dir).resolve())
        return real_stock_loader(data_dir)

    monkeypatch.setattr(BPRMF, "_read_csv_pipe", fail_historical_stock_read)
    monkeypatch.setattr(BPRMF, "_load_item_stocks", recording_stock_loader)

    with pytest.raises(
        PermissionError,
        match="synthetic historical stock read failure",
    ):
        _run_export(paths)

    assert stock_load_dirs == [(tmp_path / "synthetic-data").resolve()]
    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


def test_success_preserves_xlsx_and_csv_formats(tmp_path, monkeypatch):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)

    assert _run_export(paths) == str(paths["xlsx"])

    workbook = load_workbook(paths["xlsx"], read_only=True, data_only=True)
    try:
        assert workbook.sheetnames == ["Рекомендации"]
        rows = list(workbook["Рекомендации"].iter_rows(values_only=True))
    finally:
        workbook.close()
    assert rows[0] == (
        "MindboxID",
        "ДисконтнаяКарта",
        "Почта",
        "Телефон",
        "КодНоменклатуры_1",
        "НазваниеНоменклатуры_1",
        "Коллекция_1",
        "Коэффициент_1",
        "Конверсия_1",
        "Остаток_1",
    )
    assert rows[1][0:6] == (
        "user-1",
        "card-1",
        "user-1@example.test",
        "+7 (900) 123-45-67",
        "item-1",
        "Synthetic item",
    )
    assert rows[1][7:] == (1, 12.34, "150")

    assert _read_csv_rows(paths["csv1"]) == [
        ["CustomerID", "ProductID"],
        ["79001234567", "item-1"],
    ]
    assert _read_csv_rows(paths["csv2"]) == [
        [
            "CustomerMindboxId",
            "Quantity",
            "ProductGroupOffline1C",
            "CustomFieldKoefficient",
        ],
        ["user-1", "1", "item-1", "1,00"],
    ]
    _assert_no_export_temps(tmp_path)


@pytest.mark.parametrize(
    ("enable_csv1", "enable_csv2"),
    [(False, True), (True, False), (False, False)],
)
def test_disabled_csv_creates_neither_final_nor_temp(
    tmp_path, monkeypatch, enable_csv1, enable_csv2
):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    disabled_paths = []
    if not enable_csv1:
        disabled_paths.append(paths["csv1"])
        paths["csv1"] = None
    if not enable_csv2:
        disabled_paths.append(paths["csv2"])
        paths["csv2"] = None

    _run_export(paths)

    assert paths["xlsx"].is_file()
    if enable_csv1:
        assert paths["csv1"].is_file()
    if enable_csv2:
        assert paths["csv2"].is_file()
    for disabled_path in disabled_paths:
        assert not disabled_path.exists()
    _assert_no_export_temps(tmp_path)


def test_custom_output_directories_are_supported(tmp_path, monkeypatch):
    _prepare_synthetic_export(tmp_path, monkeypatch)
    paths = {
        "xlsx": tmp_path / "excel-output" / "custom.xlsx",
        "csv1": tmp_path / "internet-output" / "custom-internet.csv",
        "csv2": tmp_path / "mindbox-output" / "custom-mindbox.csv",
    }

    _run_export(paths)

    assert paths["xlsx"].is_file()
    assert _read_csv_rows(paths["csv1"])[0] == ["CustomerID", "ProductID"]
    assert _read_csv_rows(paths["csv2"])[0] == [
        "CustomerMindboxId",
        "Quantity",
        "ProductGroupOffline1C",
        "CustomFieldKoefficient",
    ]
    _assert_no_export_temps(tmp_path)


class _FailingCsvWriter:
    def __init__(self, writer, *, fail_on_row):
        self.writer = writer
        self.fail_on_row = fail_on_row
        self.rows_written = 0

    def writerow(self, row):
        self.rows_written += 1
        result = self.writer.writerow(row)
        if self.rows_written == self.fail_on_row:
            raise OSError("synthetic csv writer error")
        return result


@pytest.mark.parametrize(
    ("target_name", "fail_on_row"),
    [("InternetMagazin.csv", 1), ("InternetMagazin.csv", 2), ("Mindbox.csv", 1)],
)
def test_csv_write_error_keeps_all_existing_outputs(
    tmp_path, monkeypatch, target_name, fail_on_row
):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)
    real_writer = BPRMF.csv.writer

    def failing_writer(file_object, *args, **kwargs):
        writer = real_writer(file_object, *args, **kwargs)
        if target_name in Path(file_object.name).name:
            return _FailingCsvWriter(writer, fail_on_row=fail_on_row)
        return writer

    monkeypatch.setattr(BPRMF.csv, "writer", failing_writer)

    with pytest.raises(OSError, match="synthetic csv writer error"):
        _run_export(paths)

    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


class _CsvFileProxy:
    def __init__(
        self,
        file_object,
        *,
        fail_flush=False,
        fail_close=False,
        close_attempts=None,
    ):
        self.file_object = file_object
        self.fail_flush = fail_flush
        self.fail_close = fail_close
        self.close_attempts = close_attempts

    def __getattr__(self, name):
        return getattr(self.file_object, name)

    def write(self, value):
        return self.file_object.write(value)

    def flush(self):
        if self.fail_flush:
            raise OSError("synthetic csv flush error")
        return self.file_object.flush()

    def close(self):
        if self.close_attempts is not None:
            self.close_attempts.append(Path(self.file_object.name).name)
        self.file_object.close()
        if self.fail_close:
            raise OSError("synthetic csv close error")


def _patch_csv_output_open(
    monkeypatch,
    *,
    fail_flush_for=None,
    fail_close_for=None,
    close_attempts=None,
):
    real_open = open

    def wrapped_open(path, mode="r", *args, **kwargs):
        file_object = real_open(path, mode, *args, **kwargs)
        path_name = Path(path).name
        if "w" in mode and (
            "InternetMagazin.csv" in path_name or "Mindbox.csv" in path_name
        ):
            return _CsvFileProxy(
                file_object,
                fail_flush=bool(fail_flush_for and fail_flush_for in path_name),
                fail_close=bool(fail_close_for and fail_close_for in path_name),
                close_attempts=close_attempts,
            )
        return file_object

    monkeypatch.setattr(BPRMF, "open", wrapped_open, raising=False)


def test_csv_flush_error_prevents_publication(tmp_path, monkeypatch):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)
    _patch_csv_output_open(monkeypatch, fail_flush_for="InternetMagazin.csv")

    with pytest.raises(OSError, match="synthetic csv flush error"):
        _run_export(paths)

    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


def test_first_csv_close_error_still_closes_second_csv(tmp_path, monkeypatch):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)
    close_attempts = []
    _patch_csv_output_open(
        monkeypatch,
        fail_close_for="InternetMagazin.csv",
        close_attempts=close_attempts,
    )

    with pytest.raises(OSError, match="synthetic csv close error"):
        _run_export(paths)

    assert any("InternetMagazin.csv" in name for name in close_attempts)
    assert any("Mindbox.csv" in name for name in close_attempts)
    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


def test_primary_error_is_not_hidden_by_csv_close_error(
    tmp_path, monkeypatch, capsys
):
    paths = _prepare_synthetic_export(
        tmp_path,
        monkeypatch,
        model=_SyntheticExportModel(
            inference_error=RuntimeError("synthetic primary inference error")
        ),
    )
    old_bytes = _write_old_outputs(paths)
    _patch_csv_output_open(
        monkeypatch,
        fail_close_for="InternetMagazin.csv",
    )

    with pytest.raises(RuntimeError, match="synthetic primary inference error"):
        _run_export(paths)

    assert "synthetic csv close error" in capsys.readouterr().err
    _assert_old_outputs(paths, old_bytes)


class _FailingWorkbookSave:
    def __init__(self, workbook):
        self.workbook = workbook
        self.failed = False

    def __getattr__(self, name):
        return getattr(self.workbook, name)

    def save(self, path):
        if self.failed:
            return self.workbook.save(path)
        self.failed = True
        Path(path).write_bytes(b"partial-xlsx")
        raise OSError("synthetic xlsx save error")


def test_partial_xlsx_save_keeps_existing_xlsx(tmp_path, monkeypatch):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)
    real_workbook = BPRMF.Workbook
    monkeypatch.setattr(
        BPRMF,
        "Workbook",
        lambda *args, **kwargs: _FailingWorkbookSave(
            real_workbook(*args, **kwargs)
        ),
    )

    with pytest.raises(OSError, match="synthetic xlsx save error"):
        _run_export(paths)

    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


def test_xlsx_validation_error_keeps_existing_outputs(tmp_path, monkeypatch):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)

    def fail_validation(path):
        raise OSError("synthetic xlsx validation error")

    monkeypatch.setattr(
        BPRMF,
        "_validate_export_xlsx",
        fail_validation,
        raising=False,
    )

    with pytest.raises(OSError, match="synthetic xlsx validation error"):
        _run_export(paths)

    _assert_old_outputs(paths, old_bytes)
    _assert_no_export_temps(tmp_path)


@pytest.mark.parametrize("failed_replace_number", [1, 2, 3])
def test_replace_failure_preserves_file_level_invariant(
    tmp_path, monkeypatch, failed_replace_number
):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    old_bytes = _write_old_outputs(paths)
    real_replace = BPRMF.os.replace
    replace_destinations = []

    def fail_selected_replace(source, destination):
        replace_destinations.append(Path(destination))
        if len(replace_destinations) == failed_replace_number:
            raise PermissionError("synthetic publish permission error")
        real_replace(source, destination)

    monkeypatch.setattr(BPRMF.os, "replace", fail_selected_replace)

    with pytest.raises(PermissionError, match="synthetic publish permission error"):
        _run_export(paths)

    assert [path.name for path in replace_destinations] == [
        "InternetMagazin.csv",
        "Mindbox.csv",
        "Рекомендации.xlsx",
    ][:failed_replace_number]
    if failed_replace_number == 1:
        _assert_old_outputs(paths, old_bytes)
    else:
        assert _read_csv_rows(paths["csv1"])[0] == ["CustomerID", "ProductID"]
        if failed_replace_number == 2:
            _assert_old_outputs(paths, old_bytes, names=("csv2", "xlsx"))
        else:
            assert _read_csv_rows(paths["csv2"])[0][0] == "CustomerMindboxId"
            _assert_old_outputs(paths, old_bytes, names=("xlsx",))
    _assert_no_export_temps(tmp_path)


def test_cleanup_error_does_not_hide_primary_error(tmp_path, monkeypatch, capsys):
    paths = _prepare_synthetic_export(
        tmp_path,
        monkeypatch,
        model=_SyntheticExportModel(
            inference_error=RuntimeError("synthetic primary export error")
        ),
    )
    old_bytes = _write_old_outputs(paths)
    real_remove = BPRMF.os.remove

    def fail_export_temp_cleanup(path):
        if Path(path).is_relative_to(tmp_path) and Path(path).suffix == ".tmp":
            raise OSError("synthetic export temp cleanup error")
        return real_remove(path)

    monkeypatch.setattr(BPRMF.os, "remove", fail_export_temp_cleanup)

    with pytest.raises(RuntimeError, match="synthetic primary export error"):
        _run_export(paths)

    assert "synthetic export temp cleanup error" in capsys.readouterr().err
    _assert_old_outputs(paths, old_bytes)


def test_failure_cleanup_does_not_remove_another_publishers_temp(
    tmp_path, monkeypatch
):
    paths = _prepare_synthetic_export(
        tmp_path,
        monkeypatch,
        model=_SyntheticExportModel(
            inference_error=RuntimeError("synthetic inference error")
        ),
    )
    old_bytes = _write_old_outputs(paths)
    foreign_temp = paths["csv1"].parent / ".InternetMagazin.foreign.csv.tmp"
    foreign_bytes = b"another-publisher-temp"
    foreign_temp.write_bytes(foreign_bytes)

    with pytest.raises(RuntimeError, match="synthetic inference error"):
        _run_export(paths)

    assert foreign_temp.read_bytes() == foreign_bytes
    _assert_old_outputs(paths, old_bytes)


def test_parallel_exporters_use_independent_temp_paths(tmp_path, monkeypatch):
    paths = _prepare_synthetic_export(tmp_path, monkeypatch)
    real_replace = BPRMF.os.replace
    first_publication_barrier = threading.Barrier(2)
    replace_lock = threading.Lock()
    records_lock = threading.Lock()
    thread_state = threading.local()
    recorded_sources = []

    def synchronized_replace(source, destination):
        call_number = getattr(thread_state, "call_number", 0)
        thread_state.call_number = call_number + 1
        with records_lock:
            recorded_sources.append((Path(source), Path(destination)))
        if call_number == 0:
            first_publication_barrier.wait(timeout=10)
        with replace_lock:
            real_replace(source, destination)

    monkeypatch.setattr(BPRMF.os, "replace", synchronized_replace)

    def export_once():
        try:
            _run_export(paths)
        except Exception as error:
            return error
        return None

    with ThreadPoolExecutor(max_workers=2) as executor:
        errors = list(executor.map(lambda _index: export_once(), range(2)))

    assert errors == [None, None]
    assert len(recorded_sources) == 6
    assert len({source for source, _destination in recorded_sources}) == 6
    for destination in paths.values():
        matching_sources = {
            source
            for source, recorded_destination in recorded_sources
            if recorded_destination == destination
        }
        assert len(matching_sources) == 2
        assert all(source.parent == destination.parent for source in matching_sources)
    assert _read_csv_rows(paths["csv1"])[1] == ["79001234567", "item-1"]
    assert _read_csv_rows(paths["csv2"])[1][-1] == "1,00"
    load_workbook(paths["xlsx"], read_only=True).close()
    _assert_no_export_temps(tmp_path)
